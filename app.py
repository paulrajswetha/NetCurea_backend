from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from pymongo import MongoClient
import bcrypt
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import io
from datetime import datetime
import uuid
from bson import ObjectId
from pymongo import ReturnDocument
import base64
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import textwrap
import requests
import random
from bson.objectid import ObjectId
from joblib import load
import numpy as np
import sklearn 

# Placeholder Gemini AI API key (replace with your actual key)
GEMINI_API_KEY = "AIzaSyBxA-q-y9rQkP6Lfxc9kE7bMr6_AZfx-RA"

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)
logging.getLogger("pymongo").setLevel(logging.WARNING)
app = Flask(__name__)
CORS(app)
app.secret_key = 'supersecretkey'
CORS(app, resources={r"/*": {"origins": ["http://localhost:3000", "http://localhost:5173"]}})

def init_db():
    try:
        client = MongoClient("mongodb+srv://medigrid:medigrid2025@cluster0.amqrmdr.mongodb.net/")
        db = client['hospitals']
        collections = ['users', 'appointments', 'prescriptions', 'medical_records', 'availability', 'activity_log', 'bills']
        for collection in collections:
            if collection not in db.list_collection_names():
                db.create_collection(collection)
        users = db['users']
        hashed = bcrypt.hashpw('adminpass'.encode('utf-8'), bcrypt.gensalt())
        users.update_one(
            {'user_id': 'AD_root'},
            {'$setOnInsert': {
                'user_id': 'AD_root',
                'password': hashed.decode('utf-8'),
                'role': 'admin',
                'name': 'Admin',
                'email': 'admin@hospital.com',
                'image': '',
                'about': '',
                'created_at': datetime.now()
            }},
            upsert=True
        )
        return db
    except Exception as e:
        logger.error(f"Database initialization error: {str(e)}")
        raise

db = init_db()

def get_db_collection(collection_name):
    try:
        return db[collection_name]
    except Exception as e:
        logger.error(f"Database connection error: {e}")
        return None

def log_activity(action, user_id):
    try:
        activity_log = get_db_collection('activity_log')
        if activity_log is not None:
            activity_log.insert_one({
                'action': action,
                'user_id': user_id or 'anonymous',
                'timestamp': datetime.now()
            })
    except Exception as e:
        logger.error(f"Error logging activity: {str(e)}")

def generate_token_number(doctor_user_id, date):
    appointments = get_db_collection('appointments')
    tokens = list(appointments.find({'doctor_user_id': doctor_user_id, 'date': date}, {'token_number': 1, 'token_seq': 1}))
    used_numbers = []
    for t in tokens:
        if t.get('token_seq'):
            try:
                num = int(t['token_seq'])
                used_numbers.append(num)
            except Exception:
                continue
        elif t.get('token_number'):
            try:
                num = int(str(t['token_number']).split('_')[-1])
                used_numbers.append(num)
            except Exception:
                continue
    next_number = 1
    if used_numbers:
        next_number = max(used_numbers) + 1
    token_seq = f"{next_number:02d}"
    token = f"TK_{doctor_user_id}_{date}_{next_number:04d}"
    return token, token_seq

model = load('diabetes_rf_model.joblib')  # use your actual path

@app.route("/predict_risk", methods=["POST"])
def predict_risk():
    try:
        data = request.get_json()
        features = data.get("features")
        if not features:
            return jsonify({"error": "No features provided"}), 400

        prediction = model.predict([features])[0]
        proba = model.predict_proba([features])[0].tolist()

        return jsonify({
            "prediction": int(prediction),
            "confidence": round(max(proba), 4),
            "probabilities": proba
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/chatbot', methods=['POST'])
def chatbot():
    try:
        data = request.json
        user_message = data.get('message', '').strip()
        if not user_message:
            return jsonify({'message': 'Empty query provided'}), 400

        # Sanitize input to prevent injection
        if any(char in user_message for char in ['<', '>', ';', '&']):
            return jsonify({'message': 'Invalid characters in query'}), 400

        context = """
        You are a doctor/guide like chatbot for NetCurea, a hospital management system. Answer queries related to the system's functionality, including:
        - Hospital information (name, location, contact details)
        - Doctor and patient management (registration, roles, scheduling)
        - Appointments (booking, updating, cancelling, token numbers)
        - Prescriptions and medical records (creation, retrieval)
        - General platform features (e.g., secure records, notifications, billing)
        -otherwise chat normally 
        Do not reveal sensitive information such as passwords, user IDs, or personal identifiable information (PII). 
        If the query is unclear or unrelated, respond with: 
        "I'm sorry, I can only assist with questions related to the NetCurea system. Please ask about hospitals, doctors, appointments, or other platform features."
        Current date: June 29, 2025
        """

        greetings = ['hi', 'hello', 'hey', 'greetings']
        if user_message.lower() in greetings:
            response = {
                'message': "Hello! I'm the NetCurea AI Assistant. How can I help you today? You can ask about hospitals, doctors, appointments, registration, or anything else about the NetCurea platform!",
                'sender': 'bot'
            }
            log_activity('Chatbot greeting query', 'anonymous')
            return jsonify(response)

        users = get_db_collection('users')
        if users is None:
            return jsonify({'message': 'Error accessing database. Please try again later.'}), 500

        lower_message = user_message.lower()
        hospital_query = 'hospital' in lower_message or 'hospitals' in lower_message
        if hospital_query:
            query = {'role': 'hospital'}
            location = None
            if 'near me' in lower_message or 'nearby' in lower_message:
                location_query = lower_message.replace('near me', '').replace('nearby', '').strip()
                if location_query and 'hospital' in location_query:
                    location = location_query.split('hospital')[-1].strip()
                if not location:
                    response = {
                        'message': "Please provide a specific location (e.g., 'hospitals in Delhi') since geolocation is not available.",
                        'sender': 'bot'
                    }
                    log_activity('Chatbot hospital location query (no location)', 'anonymous')
                    return jsonify(response)
                query['location'] = {'$regex': location, '$options': 'i'}
            elif 'more hospitals' in lower_message or lower_message in ['hospitals', 'hospital info', 'hospital information']:
                pass
            else:
                hospital_name = lower_message.replace('hospital', '').strip()
                if hospital_name:
                    query['name'] = {'$regex': hospital_name, '$options': 'i'}

            hospitals = list(users.find(query))
            if hospitals:
                hospital_list = [
                    {
                        'name': h.get('name', 'Unknown'),
                        'location': h.get('location', 'Not available'),
                        'phone': h.get('phone', 'Not available'),
                        'email': h.get('email', 'Not available'),
                        'about': h.get('about', 'No description available')
                    } for h in hospitals
                ]
                response_message = "Here are the hospitals I found:\n" + "\n".join(
                    f"- {h['name']} (Location: {h['location']}, Contact: {h['phone']}, Email: {h['email']}, About: {h['about']})"
                    for h in hospital_list
                )
                response = {'message': response_message, 'sender': 'bot'}
                log_activity(f'Chatbot hospital query (found {len(hospitals)} hospitals)', 'anonymous')
                return jsonify(response)
            else:
                response = {
                    'message': f"No hospitals found matching your query. Try asking for a specific hospital or location, or contact support@netcurea.com.",
                    'sender': 'bot'
                }
                log_activity('Chatbot hospital query (no results)', 'anonymous')
                return jsonify(response)

        doctor_query = 'doctor' in lower_message or 'doctors' in lower_message or 'doctor info' in lower_message or 'doctor information' in lower_message
        if doctor_query:
            query = {'role': 'doctor'}
            hospital_name = None
            specialization = None
            if 'in' in lower_message and 'hospital' in lower_message:
                hospital_name = lower_message.split('hospital')[-1].strip()
                if hospital_name:
                    hospital = users.find_one({'role': 'hospital', 'name': {'$regex': hospital_name, '$options': 'i'}})
                    if hospital:
                        query['hospital_user_id'] = hospital.get('user_id')
            elif 'special' in lower_message:
                specialization = lower_message.split('special')[-1].strip()
                if specialization:
                    query['specialization'] = {'$regex': specialization, '$options': 'i'}

            doctors = list(users.find(query))
            if doctors:
                doctor_list = []
                for d in doctors:
                    hospital = users.find_one({'user_id': d.get('hospital_user_id'), 'role': 'hospital'})
                    doctor_info = {
                        'name': d.get('name', 'Unknown'),
                        'specialization': d.get('specialization', 'Not specified'),
                        'hospital': hospital.get('name', 'Not assigned') if hospital else 'Not assigned',
                        'phone': d.get('phone', 'Not available'),
                        'email': d.get('email', 'Not available')
                    }
                    doctor_list.append(doctor_info)
                response_message = "Here are the doctors I found:\n" + "\n".join(
                    f"- {d['name']} (Specialization: {d['specialization']}, Hospital: {d['hospital']}, Contact: {d['phone']}, Email: {d['email']})"
                    for d in doctor_list
                )
                response = {'message': response_message, 'sender': 'bot'}
                log_activity(f'Chatbot doctor query (found {len(doctors)} doctors)', 'anonymous')
                return jsonify(response)
            else:
                response = {
                    'message': "No doctors found matching your query. Try specifying a specialization or hospital, or contact support@netcurea.com.",
                    'sender': 'bot'
                }
                log_activity('Chatbot doctor query (no results)', 'anonymous')
                return jsonify(response)

        try:
            gemini_response = requests.post(
                'https://api.gemini.ai/v1/chat',
                headers={'Authorization': f'Bearer {GEMINI_API_KEY}'},
                json={'prompt': f"{context}\nUser query: {user_message}"}
            )
            gemini_response.raise_for_status()
            bot_response = gemini_response.json().get('response', 'Sorry, I could not process your request.')
        except Exception as e:
            logger.error(f"Gemini AI error: {str(e)}")
            if 'register as doctor' in lower_message or 'register as a doctor' in lower_message:
                bot_response = (
                    "To register as a doctor, please email the admin at support@netcurea.com with the following details:\n"
                    "• Role: Doctor\n• Name\n• Specialization\n• Phone\n• Email\n• Age\n• Gender\n\n"
                    "The admin will provide a confirmation and payment link. After payment, you will receive login credentials."
                )
            elif 'register as patient' in lower_message or 'register as a patient' in lower_message:
                bot_response = (
                    "To register as a patient, please email the admin at support@netcurea.com with the following details:\n"
                    "• Role: Patient\n• Name\n• Phone\n• Email\n• Age\n• Gender\n\n"
                    "The admin will provide a confirmation and payment link. After payment, you will receive login credentials."
                )
            elif 'register as hospital' in lower_message or 'register as a hospital' in lower_message:
                bot_response = (
                    "To register as a hospital, please email the admin at support@netcurea.com with the following details:\n"
                    "• Role: Hospital\n• Name\n• Phone\n• Email\n• Location\n\n"
                    "The admin will provide a confirmation and payment link. After payment, you will receive login credentials."
                )
            elif 'login' in lower_message:
                bot_response = (
                    "To log in, visit the login page or contact the admin at support@netcurea.com for assistance with your credentials."
                )
            else:
                bot_response = (
                    "I'm not sure how to assist with that. Please ask about hospitals, doctors, registration, or login, "
                    "or contact the admin at support@netcurea.com for further assistance."
                )

        log_activity('Chatbot query processed', 'anonymous')
        return jsonify({'message': bot_response, 'sender': 'bot'})
    except Exception as e:
        logger.error(f"Chatbot error: {str(e)}")
        return jsonify({'message': 'Error processing your request. Please try again.'}), 500

@app.route('/recent_activity', methods=['GET'])
def recent_activity():
    try:
        limit = request.args.get('limit')
        activity_log = get_db_collection('activity_log')
        if activity_log is not None:
            query = activity_log.find().sort('timestamp', -1)
            if limit is not None:
                try:
                    limit = int(limit)
                    query = query.limit(limit)
                except ValueError:
                    pass
            activities = list(query)
            for activity in activities:
                activity['_id'] = str(activity['_id'])
                activity['timestamp'] = activity['timestamp'].isoformat()
            return jsonify(activities)
        else:
            return jsonify({'message': 'Failed to access activity log'}), 500
    except Exception as e:
        logger.error(f"Error fetching recent activity: {str(e)}")
        return jsonify({'message': 'Error fetching recent activity'}), 500

@app.route('/login', methods=['POST'])
def login():
    data = request.json
    user_id = data.get('user_id')
    password = data.get('password')
    users = get_db_collection('users')
    user = users.find_one({'user_id': user_id})
    if user and bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):
        user_data = {
            'message': 'Login successful',
            'role': user['role'],
            'name': user['name'],
            'user_id': user['user_id'],
            'hospital_user_id': user.get('hospital_user_id', ''),
            'specialization': user.get('specialization', '')
        }
        log_activity(f"User logged in: {user_id}", user_id)
        return jsonify(user_data)
    return jsonify({'message': 'Invalid user ID or password'}), 401

@app.route('/export/<type>/excel', methods=['GET'])
def export_excel(type):
    try:
        query = {}
        if request.args.get('doctor_user_id'):
            query['doctor_user_id'] = request.args.get('doctor_user_id')
        if request.args.get('patient_user_id'):
            query['patient_user_id'] = request.args.get('patient_user_id')

        wb = Workbook()
        ws = wb.active

        if type == 'doctors':
            collection = get_db_collection('users')
            query['role'] = 'doctor'
            data = list(collection.find(query))
            ws.title = "Doctors"
            headers = ['User ID', 'Name', 'Specialization', 'Phone', 'Email', 'Age', 'Gender', 'Hospital User ID']
            ws.append(headers)
            for item in data:
                ws.append([
                    item.get('user_id', ''),
                    item.get('name', ''),
                    item.get('specialization', ''),
                    item.get('phone', ''),
                    item.get('email', ''),
                    str(item.get('age', '')),
                    item.get('gender', ''),
                    item.get('hospital_user_id', '')
                ])

        elif type == 'patients':
            collection = get_db_collection('users')
            query['role'] = 'patient'
            data = list(collection.find(query))
            ws.title = "Patients"
            headers = ['User ID', 'Name', 'Phone', 'Email', 'Age', 'Gender', 'Hospital User ID']
            ws.append(headers)
            for item in data:
                ws.append([
                    item.get('user_id', ''),
                    item.get('name', ''),
                    item.get('phone', ''),
                    item.get('email', ''),
                    str(item.get('age', '')),
                    item.get('gender', ''),
                    item.get('hospital_user_id', '')
                ])

        elif type == 'hospitals':
            collection = get_db_collection('users')
            query['role'] = 'hospital'
            data = list(collection.find(query))
            ws.title = "Hospitals"
            headers = ['User ID', 'Name', 'Phone', 'Email', 'Location']
            ws.append(headers)
            for item in data:
                ws.append([
                    item.get('user_id', ''),
                    item.get('name', ''),
                    item.get('phone', ''),
                    item.get('email', ''),
                    item.get('location', '')
                ])

        elif type == 'appointments':
            collection = get_db_collection('appointments')
            data = list(collection.find(query))
            ws.title = "Appointments"
            headers = ['Token Number', 'Doctor User ID', 'Patient User ID', 'Date', 'Time', 'Status', 'Notes']
            ws.append(headers)
            for item in data:
                ws.append([
                    item.get('token_number', ''),
                    item.get('doctor_user_id', ''),
                    item.get('patient_user_id', ''),
                    item.get('date', ''),
                    item.get('time', ''),
                    item.get('status', ''),
                    item.get('notes', '')
                ])

        elif type == 'prescriptions':
            collection = get_db_collection('prescriptions')
            data = list(collection.find(query))
            ws.title = "Prescriptions"
            headers = ['Doctor User ID', 'Patient User ID', 'Token Number', 'Medication', 'Dosage', 'Frequency', 'Frequency Unit', 'Duration', 'Duration Unit', 'Instructions', 'Follow-Up', 'Date']
            ws.append(headers)
            for item in data:
                # Fetch token_number from appointment
                appointment = get_db_collection('appointments').find_one({'doctor_user_id': item.get('doctor_user_id'), 'patient_user_id': item.get('patient_user_id'), 'date': item.get('date')})
                token_number = appointment.get('token_number', '') if appointment else ''
                ws.append([
                    item.get('doctor_user_id', ''),
                    item.get('patient_user_id', ''),
                    token_number,
                    item.get('medication', ''),
                    item.get('dosage', ''),
                    item.get('frequency', ''),
                    item.get('frequency_unit', ''),
                    item.get('duration', ''),
                    item.get('duration_unit', ''),
                    item.get('instructions', ''),
                    'Yes' if item.get('follow_up', False) else 'No',
                    item.get('date', '')
                ])

        elif type == 'medical_records':
            collection = get_db_collection('medical_records')
            data = list(collection.find(query))
            ws.title = "Medical Records"
            headers = ['Doctor User ID', 'Patient User ID', 'Token Number', 'Diagnosis', 'Treatment', 'Notes', 'Date']
            ws.append(headers)
            for item in data:
                # Fetch token_number from appointment
                appointment = get_db_collection('appointments').find_one({'doctor_user_id': item.get('doctor_user_id'), 'patient_user_id': item.get('patient_user_id'), 'date': item.get('date')})
                token_number = appointment.get('token_number', '') if appointment else ''
                ws.append([
                    item.get('doctor_user_id', ''),
                    item.get('patient_user_id', ''),
                    token_number,
                    item.get('diagnosis', ''),
                    item.get('treatment', ''),
                    item.get('notes', ''),
                    item.get('date', '')
                ])

        elif type == 'bills':
            collection = get_db_collection('bills')
            data = list(collection.find(query))
            ws.title = "Bills"
            headers = ['Appointment ID', 'Patient User ID', 'Token Number', 'Amount', 'Status', 'Date']
            ws.append(headers)
            for item in data:
                # Fetch token_number from appointment
                appointment = get_db_collection('appointments').find_one({'_id': ObjectId(item.get('appointment_id'))}) if item.get('appointment_id') else None
                token_number = appointment.get('token_number', '') if appointment else ''
                ws.append([
                    item.get('appointment_id', ''),
                    item.get('patient_user_id', ''),
                    token_number,
                    str(item.get('amount', '')),
                    item.get('status', ''),
                    item.get('date', '')
                ])

        else:
            return jsonify({'message': 'Invalid export type'}), 400

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{type}_export.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}")
        return jsonify({'message': 'Error generating Excel file'}), 500

@app.route('/users', methods=['GET', 'POST'])
def handle_users():
    users = get_db_collection('users')
    if request.method == 'GET':
        role = request.args.get('role', '')
        name = request.args.get('name', '')
        specialization = request.args.get('specialization', '')
        query = {}
        if role:
            query['role'] = role
        if name:
            query['name'] = {'$regex': name, '$options': 'i'}
        if specialization and role == 'doctor':
            query['specialization'] = {'$regex': specialization, '$options': 'i'}
        user_list = list(users.find(query))
        for user in user_list:
            user['_id'] = str(user['_id'])
            user.pop('password', None)
        return jsonify(user_list)
    elif request.method == 'POST':
        data = request.json
        role = data.get('role')
        password = data.get('password')
        name = data.get('name', '')
        specialization = data.get('specialization', '')
        phone = data.get('phone', '')
        email = data.get('email', '')
        age = data.get('age')
        gender = data.get('gender', '')
        image = data.get('image', '')
        about = data.get('about', '')
        location = data.get('location', '')
        hospital_user_id = data.get('hospital_user_id', '')

        if role == 'admin':
            return jsonify({'message': 'Cannot create admin users'}), 403
        if role in ['doctor', 'patient'] and (not hospital_user_id or not hospital_user_id.startswith('HP_')):
            return jsonify({'message': 'Valid hospital_user_id is required for doctors and patients'}), 400

        try:
            user_id = generate_user_id(role, users)
        except ValueError:
            return jsonify({'message': 'Invalid role'}), 400
        
        if users.find_one({'user_id': user_id}):
            return jsonify({'message': 'User ID already exists'}), 400
        hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
        user_data = {
            'user_id': user_id,
            'password': hashed.decode('utf-8'),
            'role': role,
            'name': name,
            'specialization': specialization,
            'phone': phone,
            'email': email,
            'age': age,
            'gender': gender,
            'image': image,
            'about': about
        }
        if role in ['doctor', 'patient']:
            user_data['hospital_user_id'] = hospital_user_id
        elif role == 'hospital':
            user_data['location'] = location

        users.insert_one(user_data)
        log_activity(f"User created: {user_id}", user_id)
        return jsonify({'message': 'User added successfully', 'user_id': user_id})

@app.route('/users/<user_id>', methods=['GET', 'PUT', 'DELETE'])
def handle_user(user_id):
    users = get_db_collection('users')
    user = users.find_one({'user_id': user_id})
    if not user:
        return jsonify({'message': 'User not found'}), 404
    
    if request.method == 'GET':
        user['_id'] = str(user['_id'])
        user.pop('password', None)
        if user.get('role') != 'admin' and user.get('hospital_user_id'):
            hospital = users.find_one({'user_id': user['hospital_user_id'], 'role': 'hospital'})
            user['hospital_name'] = hospital['name'] if hospital else 'Not Assigned'
        else:
            user['hospital_name'] = None
        return jsonify(user)
    
    if request.method == 'PUT':
        data = request.json
        password = data.get('password')
        role = data.get('role', user['role'])
        name = data.get('name', user['name'])
        specialization = data.get('specialization', user['specialization']) if user.get('role') == 'doctor' else user.get('specialization', '')
        phone = data.get('phone', user['phone'])
        email = data.get('email', user['email'])
        age = data.get('age', user['age'])
        gender = data.get('gender', user['gender'])
        image = data.get('image', user.get('image', ''))
        about = data.get('about', user.get('about', ''))
        location = data.get('location', user.get('location', ''))
        hospital_user_id = data.get('hospital_user_id', user.get('hospital_user_id', ''))

        if role == 'admin':
            return jsonify({'message': 'Cannot update to admin role'}), 403
        if role in ['doctor', 'patient'] and (not hospital_user_id or not hospital_user_id.startswith('HP_')):
            return jsonify({'message': 'Valid hospital_user_id is required for doctors and patients'}), 400
        if image and image.startswith('data:image'):
            try:
                header, encoded = image.split(',', 1)
                img_data = base64.b64decode(encoded)
                if len(img_data) > 5 * 1024 * 1024:
                    return jsonify({'message': 'Image size exceeds 5MB'}), 400
                if not (header.startswith('data:image/png') or header.startswith('data:image/jpeg')):
                    return jsonify({'message': 'Only PNG and JPEG images are supported'}), 400
            except Exception as e:
                return jsonify({'message': f'Invalid image data: {str(e)}'}), 400
        update_data = {
            'role': role,
            'name': name,
            'specialization': specialization,
            'phone': phone,
            'email': email,
            'age': age,
            'gender': gender,
            'image': image,
            'about': about
        }
        if role in ['doctor', 'patient']:
            update_data['hospital_user_id'] = hospital_user_id
        elif role == 'hospital':
            update_data['location'] = location
        if password:
            update_data['password'] = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        users.update_one({'user_id': user_id}, {'$set': update_data})
        log_activity(f"User updated: {user_id}", user_id)
        return jsonify({'message': 'User updated successfully'})
    
    elif request.method == 'DELETE':
        users.delete_one({'user_id': user_id})
        get_db_collection('appointments').delete_many({
            '$or': [
                {'doctor_user_id': user_id},
                {'patient_user_id': user_id}
            ]
        })
        log_activity(f"User deleted: {user_id}", user_id)
        return jsonify({'message': 'User deleted successfully'})

@app.route('/appointments', methods=['GET', 'POST'])
def handle_appointments():
    appointments = get_db_collection('appointments')
    if request.method == 'GET':
        doctor_user_id = request.args.get('doctor_user_id')
        patient_user_id = request.args.get('patient_user_id')
        hospital_user_id = request.args.get('hospital_user_id')
        query = {}
        if doctor_user_id:
            query['doctor_user_id'] = doctor_user_id
        if patient_user_id:
            query['patient_user_id'] = patient_user_id
        if hospital_user_id:
            users = get_db_collection('users')
            doctors = users.find({'role': 'doctor', 'hospital_user_id': hospital_user_id})
            doctor_ids = [doc['user_id'] for doc in doctors]
            query['doctor_user_id'] = {'$in': doctor_ids}
        appt_list = list(appointments.find(query))
        for appt in appt_list:
            appt['_id'] = str(appt['_id'])
            patient = get_db_collection('users').find_one({'user_id': appt['patient_user_id']})
            appt['patient_name'] = patient['name'] if patient else 'Unknown'
            doctor = get_db_collection('users').find_one({'user_id': appt['doctor_user_id']})
            appt['doctor_name'] = doctor['name'] if doctor else 'Unknown'
        return jsonify(appt_list)
    elif request.method == 'POST':
        data = request.json
        doctor_user_id = data.get('doctor_user_id')
        patient_user_id = data.get('patient_user_id')
        date = data.get('date')
        time = data.get('time')
        notes = data.get('notes', '')
        amount = data.get('amount', 1000)  # Default amount for appointment
        availability = get_db_collection('availability')
        logger.debug(f"Checking availability for doctor: {doctor_user_id}, date: {date}, time: {time}")
        avail = availability.find_one({'doctor_user_id': doctor_user_id, 'date': date, 'time': time})
        if not avail:
            logger.debug(f"No availability found in collection: {list(availability.find())}")
            return jsonify({'message': 'Time slot not available'}), 400
        existing = appointments.find_one({'doctor_user_id': doctor_user_id, 'date': date, 'time': time})
        if existing:
            logger.debug(f"Existing appointment found: {existing}")
            return jsonify({'message': 'Time slot already booked'}), 400
        token_number, token_seq = generate_token_number(doctor_user_id, date)
        appointment_id = appointments.insert_one({
            'doctor_user_id': doctor_user_id,
            'patient_user_id': patient_user_id,
            'date': date,
            'time': time,
            'status': 'Scheduled',
            'notes': notes,
            'payment_status': 'Pending',
            'token_number': token_number,
            'token_seq': token_seq,
            'amount': amount
        }).inserted_id
        availability.delete_one({'doctor_user_id': doctor_user_id, 'date': date, 'time': time})
        bills = get_db_collection('bills')
        bills.insert_one({
            'appointment_id': str(appointment_id),
            'patient_user_id': patient_user_id,
            'amount': amount,
            'status': 'Pending',
            'date': datetime.now().strftime('%Y-%m-%d')
        })
        logger.debug(f"Appointment booked successfully with ID: {appointment_id}, Token: {token_seq}")
        log_activity(f"Appointment booked: {appointment_id}, Token: {token_seq}", patient_user_id)
        return jsonify({
            'message': 'Appointment booked successfully',
            '_id': str(appointment_id),
            'payment_status': 'Pending',
            'token_number': token_seq
        })

@app.route('/appointments/<id>', methods=['PUT', 'DELETE'])
def handle_appointment(id):
    appointments = get_db_collection('appointments')
    availability = get_db_collection('availability')
    try:
        obj_id = ObjectId(id)
    except:
        return jsonify({'message': 'Invalid appointment ID format'}), 400

    if request.method == 'PUT':
        data = request.json
        update_data = {
            'doctor_user_id': data.get('doctor_user_id'),
            'patient_user_id': data.get('patient_user_id'),
            'date': data.get('date'),
            'time': data.get('time'),
            'notes': data.get('notes', ''),
            'status': data.get('status', 'Scheduled'),
            'amount': data.get('amount')
        }
        update_data = {k: v for k, v in update_data.items() if v is not None}
        result = appointments.find_one_and_update(
            {'_id': obj_id},
            {'$set': update_data},
            return_document=ReturnDocument.AFTER
        )
        if not result:
            return jsonify({'message': 'Appointment not found'}), 404
        if 'amount' in update_data:
            get_db_collection('bills').update_one(
                {'appointment_id': str(obj_id)},
                {'$set': {'amount': update_data['amount'], 'date': datetime.now().strftime('%Y-%m-%d')}}
            )
        result['_id'] = str(result['_id'])
        patient = get_db_collection('users').find_one({'user_id': result['patient_user_id']})
        result['patient_name'] = patient['name'] if patient else 'Unknown'
        doctor = get_db_collection('users').find_one({'user_id': result['doctor_user_id']})
        result['doctor_name'] = doctor['name'] if doctor else 'Unknown'
        log_activity(f"Appointment updated: {id}", result.get('patient_user_id'))
        return jsonify({'message': 'Appointment updated successfully', 'appointment': result})
    elif request.method == 'DELETE':
        try:
            appointment = appointments.find_one({'_id': obj_id})
            if not appointment:
                return jsonify({'message': 'Appointment not found'}), 404

            doctor_user_id = appointment['doctor_user_id']
            date = appointment['date']
            time = appointment['time']

            result = appointments.delete_one({'_id': obj_id})
            if result.deleted_count == 0:
                return jsonify({'message': 'Failed to delete appointment'}), 400

            existing_slot = availability.find_one({
                'doctor_user_id': doctor_user_id,
                'date': date,
                'time': time
            })

            if not existing_slot:
                availability.insert_one({
                    'doctor_user_id': doctor_user_id,
                    'date': date,
                    'time': time
                })
                logger.debug(f"Availability restored for doctor: {doctor_user_id}, date: {date}, time: {time}")

            get_db_collection('bills').delete_one({'appointment_id': str(obj_id)})
            log_activity(f"Appointment cancelled: {id}", appointment.get('patient_user_id'))
            return jsonify({'message': 'Appointment cancelled successfully'})
        except Exception as e:
            logger.error(f"Error cancelling appointment: {str(e)}")
            return jsonify({'message': f'Error cancelling appointment: {str(e)}'}), 500

@app.route('/prescriptions', methods=['GET', 'POST'])
def handle_prescriptions():
    prescriptions = get_db_collection('prescriptions')
    if request.method == 'GET':
        doctor_user_id = request.args.get('doctor_user_id')
        patient_user_id = request.args.get('patient_user_id')
        query = {}
        if doctor_user_id:
            query['doctor_user_id'] = doctor_user_id
        elif patient_user_id:
            query['patient_user_id'] = patient_user_id
        prescription_list = list(prescriptions.find(query))
        for p in prescription_list:
            p['_id'] = str(p['_id'])
        return jsonify(prescription_list)
    elif request.method == 'POST':
        data = request.json
        doctor_user_id = data.get('doctor_user_id')
        patient_user_id = data.get('patient_user_id')
        medication = data.get('medication')
        dosage = data.get('dosage')
        duration = data.get('duration')
        duration_unit = data.get('duration_unit', 'days')
        frequency = data.get('frequency')
        frequency_unit = data.get('frequency_unit', 'times per day')
        instructions = data.get('instructions', '')
        follow_up = data.get('follow_up', False)
        date = datetime.now().strftime('%Y-%m-%d')
        if not all([doctor_user_id, patient_user_id, medication, dosage]):
            return jsonify({'message': 'Missing required fields'}), 400
        prescription_id = str(uuid.uuid4())
        prescriptions.insert_one({
            '_id': prescription_id,
            'doctor_user_id': doctor_user_id,
            'patient_user_id': patient_user_id,
            'medication': medication,
            'dosage': dosage,
            'duration': duration,
            'duration_unit': duration_unit,
            'frequency': frequency,
            'frequency_unit': frequency_unit,
            'instructions': instructions,
            'follow_up': follow_up,
            'date': date
        })
        log_activity(f"Prescription added: {prescription_id}", patient_user_id)
        return jsonify({'message': 'Prescription added successfully', '_id': prescription_id})

@app.route('/prescriptions/<id>', methods=['PUT', 'DELETE'])
def handle_prescription(id):
    prescriptions = get_db_collection('prescriptions')
    if request.method == 'PUT':
        data = request.json
        update_data = {
            'doctor_user_id': data.get('doctor_user_id'),
            'patient_user_id': data.get('patient_user_id'),
            'medication': data.get('medication'),
            'dosage': data.get('dosage'),
            'duration': data.get('duration'),
            'duration_unit': data.get('duration_unit', 'days'),
            'frequency': data.get('frequency'),
            'frequency_unit': data.get('frequency_unit', 'times per day'),
            'instructions': data.get('instructions', ''),
            'follow_up': data.get('follow_up', False),
            'date': datetime.now().strftime('%Y-%m-%d')
        }
        update_data = {k: v for k, v in update_data.items() if v is not None}
        result = prescriptions.find_one_and_update(
            {'_id': id},
            {'$set': update_data},
            return_document=ReturnDocument.AFTER
        )
        if not result:
            return jsonify({'message': 'Prescription not found'}), 404
        result['_id'] = str(result['_id'])
        log_activity(f"Prescription updated: {id}", result.get('patient_user_id'))
        return jsonify({'message': 'Prescription updated successfully', 'prescription': result})
    elif request.method == 'DELETE':
        prescription = prescriptions.find_one({'_id': id})
        if not prescription:
            return jsonify({'message': 'Prescription not found'}), 404
        result = prescriptions.delete_one({'_id': id})
        if result.deleted_count == 0:
            return jsonify({'message': 'Prescription not found'}), 404
        log_activity(f"Prescription deleted: {id}", prescription.get('patient_user_id'))
        return jsonify({'message': 'Prescription deleted successfully'})

@app.route('/medical_records', methods=['GET', 'POST'])
def handle_records():
    medical_records = get_db_collection('medical_records')
    if request.method == 'GET':
        doctor_user_id = request.args.get('doctor_user_id')
        patient_user_id = request.args.get('patient_user_id')
        query = {}
        if doctor_user_id:
            query['doctor_user_id'] = doctor_user_id
        elif patient_user_id:
            query['patient_user_id'] = patient_user_id
        record_list = list(medical_records.find(query))
        for r in record_list:
            r['_id'] = str(r['_id'])
        return jsonify(record_list)
    elif request.method == 'POST':
        data = request.json
        doctor_user_id = data.get('doctor_user_id')
        patient_user_id = data.get('patient_user_id')
        diagnosis = data.get('diagnosis')
        treatment = data.get('treatment')
        notes = data.get('notes', '')
        date = datetime.now().strftime('%Y-%m-%d')
        if not all([doctor_user_id, patient_user_id, diagnosis, treatment]):
            return jsonify({'message': 'Missing required fields'}), 400
        record_id = str(uuid.uuid4())
        medical_records.insert_one({
            '_id': record_id,
            'doctor_user_id': doctor_user_id,
            'patient_user_id': patient_user_id,
            'diagnosis': diagnosis,
            'treatment': treatment,
            'notes': notes,
            'date': date
        })
        log_activity(f"Medical record added: {record_id}", patient_user_id)
        return jsonify({'message': 'Medical record added successfully', '_id': record_id})

@app.route('/medical_records/<id>', methods=['PUT', 'DELETE'])
def handle_medical_record(id):
    medical_records = get_db_collection('medical_records')
    if request.method == 'PUT':
        data = request.json
        update_data = {
            'doctor_user_id': data.get('doctor_user_id'),
            'patient_user_id': data.get('patient_user_id'),
            'diagnosis': data.get('diagnosis'),
            'treatment': data.get('treatment'),
            'notes': data.get('notes', ''),
            'date': datetime.now().strftime('%Y-%m-%d')
        }
        update_data = {k: v for k, v in update_data.items() if v is not None}
        result = medical_records.find_one_and_update(
            {'_id': id},
            {'$set': update_data},
            return_document=ReturnDocument.AFTER
        )
        if not result:
            return jsonify({'message': 'Medical record not found'}), 404
        result['_id'] = str(result['_id'])
        log_activity(f"Medical record updated: {id}", result.get('patient_user_id'))
        return jsonify({'message': 'Medical record updated successfully', 'record': result})
    elif request.method == 'DELETE':
        record = medical_records.find_one({'_id': id})
        if not record:
            return jsonify({'message': 'Medical record not found'}), 404
        result = medical_records.delete_one({'_id': id})
        if result.deleted_count == 0:
            return jsonify({'message': 'Medical record not found'}), 404
        log_activity(f"Medical record deleted: {id}", record.get('patient_user_id'))
        return jsonify({'message': 'Medical record deleted successfully'})

@app.route('/availability', methods=['GET', 'POST'])
def handle_availability():
    availability = get_db_collection('availability')
    if request.method == 'GET':
        doctor_user_id = request.args.get('doctor_user_id')
        query = {}
        if doctor_user_id:
            query['doctor_user_id'] = doctor_user_id
        availability_list = list(availability.find(query))
        for a in availability_list:
            a['_id'] = str(a['_id'])
        return jsonify(availability_list)
    elif request.method == 'POST':
        data = request.json
        doctor_user_id = data.get('doctor_user_id')
        date = data.get('date')
        time = data.get('time')
        availability.insert_one({
            'doctor_user_id': doctor_user_id,
            'date': date,
            'time': time
        })
        log_activity(f"Availability added for doctor: {doctor_user_id}", doctor_user_id)
        return jsonify({'message': 'Availability added successfully'})

@app.route('/scheduler', methods=['GET', 'POST'])
def handle_scheduler():
    scheduler = get_db_collection('scheduler')
    if request.method == 'GET':
        doctor_user_id = request.args.get('doctor_user_id')
        query = {}
        if doctor_user_id:
            query['doctor_user_id'] = doctor_user_id
        schedule_list = list(scheduler.find(query))
        for s in schedule_list:
            s['_id'] = str(s['_id'])
        return jsonify(schedule_list)
    elif request.method == 'POST':
        data = request.json
        doctor_user_id = data.get('doctor_user_id')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        day = data.get('day')
        scheduler.insert_one({
            'doctor_user_id': doctor_user_id,
            'day': day,
            'start_time': start_time,
            'end_time': end_time
        })
        log_activity(f"Scheduler entry added for doctor: {doctor_user_id}", doctor_user_id)
        return jsonify({'message': 'Scheduler entry added successfully'})

@app.route('/dashboard/stats', methods=['GET'])
def dashboard_stats():
    users = get_db_collection('users')
    appointments = get_db_collection('appointments')
    prescriptions = get_db_collection('prescriptions')
    bills = get_db_collection('bills')
    patient_user_id = request.args.get('patient_user_id')
    hospital_user_id = request.args.get('hospital_user_id')
    
    total_doctors = users.count_documents({'role': 'doctor'})
    total_patients = users.count_documents({'role': 'patient'})
    total_appointments = appointments.count_documents({})
    total_prescriptions = prescriptions.count_documents({})
    
    stats = {
        'total_doctors': total_doctors,
        'total_patients': total_patients,
        'total_appointments': total_appointments,
        'total_prescriptions': total_prescriptions
    }
    
    if patient_user_id:
        total_expenses = sum(bill['amount'] for bill in bills.find({'patient_user_id': patient_user_id, 'status': 'Paid'}))
        stats['total_expenses'] = total_expenses
    
    if hospital_user_id:
        doctors = users.find({'role': 'doctor', 'hospital_user_id': hospital_user_id})
        doctor_ids = [doc['user_id'] for doc in doctors]
        hospital_appointments = list(appointments.find({'doctor_user_id': {'$in': doctor_ids}}))
        for appt in hospital_appointments:
            appt['_id'] = str(appt['_id'])
            patient = users.find_one({'user_id': appt['patient_user_id']})
            appt['patient_name'] = patient['name'] if patient else 'Unknown'
            doctor = users.find_one({'user_id': appt['doctor_user_id']})
            appt['doctor_name'] = doctor['name'] if doctor else 'Unknown'
        stats['hospital_appointments'] = hospital_appointments
    
    return jsonify(stats)

@app.route('/availability/<id>', methods=['DELETE'])
def handle_availability_delete(id):
    availability = get_db_collection('availability')
    doctor_user_id = request.args.get('doctor_user_id')
    try:
        obj_id = ObjectId(id)
    except:
        return jsonify({'message': 'Invalid availability ID format'}), 400
    result = availability.delete_one({'_id': obj_id, 'doctor_user_id': doctor_user_id})
    if result.deleted_count == 0:
        return jsonify({'message': 'Availability slot not found or not authorized'}), 404
    log_activity(f"Availability slot deleted: {id}", doctor_user_id)
    return jsonify({'message': 'Availability slot deleted successfully'})

@app.route('/download/prescription/<id>', methods=['GET'])
def download_prescription(id):
    try:
        prescriptions = get_db_collection('prescriptions')
        users = get_db_collection('users')
        prescription = prescriptions.find_one({'_id': id})
        if not prescription:
            return jsonify({'message': 'Prescription not found'}), 404

        # Fetch appointment for token and hospital/doctor info
        appointment = get_db_collection('appointments').find_one({'doctor_user_id': prescription.get('doctor_user_id'), 'patient_user_id': prescription.get('patient_user_id'), 'date': prescription.get('date')})
        token_number = appointment.get('token_number', '') if appointment else ''
        doctor = users.find_one({'user_id': prescription.get('doctor_user_id')})
        patient = users.find_one({'user_id': prescription.get('patient_user_id')})
        hospital = users.find_one({'user_id': patient.get('hospital_user_id'), 'role': 'hospital'}) if patient else None

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        # Header
        p.setFont("Helvetica-Bold", 18)
        p.drawString(50, height - 50, hospital.get('name', 'Hospital') if hospital else 'Hospital')
        p.setFont("Helvetica", 10)
        p.drawString(50, height - 65, hospital.get('location', ''))
        p.drawString(50, height - 80, f"Contact: {hospital.get('phone', '')} | Email: {hospital.get('email', '')}")
        p.line(50, height - 90, width - 50, height - 90)

        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 120, "Medical Prescription")

        y = height - 160
        p.setFont("Helvetica", 12)
        p.drawString(50, y, f"Patient: {prescription.get('patient_user_id', '')}")
        y -= 20
        p.drawString(50, y, f"Doctor: {doctor.get('name', prescription.get('doctor_user_id', '')) if doctor else prescription.get('doctor_user_id', '')}")
        y -= 20
        p.drawString(50, y, f"Date: {prescription.get('date', '')}")
        y -= 40

        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y, "Prescription Details:")
        y -= 25
        p.setFont("Helvetica", 12)
        p.drawString(50, y, f"Medication: {prescription.get('medication', '')}")
        y -= 20
        p.drawString(50, y, f"Dosage: {prescription.get('dosage', '')}")
        y -= 20
        p.drawString(50, y, f"Frequency: {prescription.get('frequency', '')} {prescription.get('frequency_unit', '')}")
        y -= 20
        p.drawString(50, y, f"Duration: {prescription.get('duration', '')} {prescription.get('duration_unit', '')}")
        y -= 20
        p.drawString(50, y, f"Instructions: {prescription.get('instructions', '')}")
        y -= 20
        p.drawString(50, y, f"Follow-up Required: {'Yes' if prescription.get('follow_up', False) else 'No'}")
        y -= 20
        p.drawString(50, y, f"Token Number: {token_number}")
        y -= 20

        # Footer
        p.line(50, 120, width - 50, 120)
        p.setFont("Helvetica", 10)
        p.drawString(50, 105, hospital.get('name', 'Hospital') + ' | ' + hospital.get('location', '') if hospital else '')
        p.setFont("Helvetica", 12)
        sign_x = width - 250
        sign_y = 135
        p.drawString(sign_x, sign_y, f"Doctor's Signature: {doctor.get('name', '') if doctor else ''}")
        p.line(sign_x, sign_y - 5, width - 50, sign_y - 5)

        p.showPage()
        p.save()
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"prescription_{id}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        logger.error(f"Error generating prescription PDF: {str(e)}")
        return jsonify({'message': 'Error generating prescription'}), 500

@app.route('/download/medical_record/<id>', methods=['GET'])
def download_medical_record(id):
    try:
        medical_records = get_db_collection('medical_records')
        users = get_db_collection('users')
        record = medical_records.find_one({'_id': id})
        if not record:
            return jsonify({'message': 'Medical record not found'}), 404

        # Fetch appointment for token and hospital/doctor info
        appointment = get_db_collection('appointments').find_one({'doctor_user_id': record.get('doctor_user_id'), 'patient_user_id': record.get('patient_user_id'), 'date': record.get('date')})
        token_number = appointment.get('token_number', '') if appointment else ''
        doctor = users.find_one({'user_id': record.get('doctor_user_id')})
        patient = users.find_one({'user_id': record.get('patient_user_id')})
        hospital = users.find_one({'user_id': patient.get('hospital_user_id'), 'role': 'hospital'}) if patient else None

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        # Header
        p.setFont("Helvetica-Bold", 18)
        p.drawString(50, height - 50, hospital.get('name', 'Hospital') if hospital else 'Hospital')
        p.setFont("Helvetica", 10)
        p.drawString(50, height - 65, hospital.get('location', ''))
        p.drawString(50, height - 80, f"Contact: {hospital.get('phone', '')} | Email: {hospital.get('email', '')}")
        p.line(50, height - 90, width - 50, height - 90)

        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 120, "Medical Record")

        y = height - 160
        p.setFont("Helvetica", 12)
        p.drawString(50, y, f"Patient: {record.get('patient_user_id', '')}")
        y -= 20
        p.drawString(50, y, f"Doctor: {doctor.get('name', record.get('doctor_user_id', '')) if doctor else record.get('doctor_user_id', '')}")
        y -= 20
        p.drawString(50, y, f"Date: {record.get('date', '')}")
        y -= 40

        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y, "Diagnosis:")
        y -= 25
        p.setFont("Helvetica", 12)
        for line in textwrap.wrap(record.get('diagnosis', ''), width=90):
            p.drawString(50, y, line)
            y -= 18
        y -= 20

        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y, "Treatment:")
        y -= 25
        p.setFont("Helvetica", 12)
        for line in textwrap.wrap(record.get('treatment', ''), width=90):
            p.drawString(50, y, line)
            y -= 18

        if record.get('notes', ''):
            y -= 20
            p.setFont("Helvetica-Bold", 14)
            p.drawString(50, y, "Additional Notes:")
            y -= 25
            p.setFont("Helvetica", 12)
            for line in textwrap.wrap(record.get('notes', ''), width=90):
                p.drawString(50, y, line)
                y -= 18
        y -= 20
        p.drawString(50, y, f"Token Number: {token_number}")
        y -= 20

        # Footer
        p.line(50, 120, width - 50, 120)
        p.setFont("Helvetica", 10)
        p.drawString(50, 105, hospital.get('name', 'Hospital') + ' | ' + hospital.get('location', '') if hospital else '')
        p.setFont("Helvetica", 12)
        sign_x = width - 250
        sign_y = 135
        p.drawString(sign_x, sign_y, f"Doctor's Signature: {doctor.get('name', '') if doctor else ''}")
        p.line(sign_x, sign_y - 5, width - 50, sign_y - 5)

        p.showPage()
        p.save()
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"medical_record_{id}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        logger.error(f"Error generating medical record PDF: {str(e)}")
        return jsonify({'message': 'Error generating medical record'}), 500

@app.route('/download/bill/<appointment_id>', methods=['GET'])
def download_bill(appointment_id):
    try:
        appointments = get_db_collection('appointments')
        bills = get_db_collection('bills')
        users = get_db_collection('users')
        appointment = appointments.find_one({'_id': ObjectId(appointment_id)})
        if not appointment:
            return jsonify({'message': 'Appointment not found'}), 404
        bill = bills.find_one({'appointment_id': appointment_id})
        if not bill:
            return jsonify({'message': 'Bill not found'}), 404
        patient = users.find_one({'user_id': appointment['patient_user_id']})
        doctor = users.find_one({'user_id': appointment['doctor_user_id']})
        hospital = users.find_one({'user_id': patient.get('hospital_user_id'), 'role': 'hospital'}) if patient else None

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        # Header
        p.setFont("Helvetica-Bold", 18)
        p.drawString(50, height - 50, hospital.get('name', 'Hospital') if hospital else 'Hospital')
        p.setFont("Helvetica", 10)
        p.drawString(50, height - 65, hospital.get('location', ''))
        p.drawString(50, height - 80, f"Contact: {hospital.get('phone', '')} | Email: {hospital.get('email', '')}")
        p.line(50, height - 90, width - 50, height - 90)

        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 120, "Invoice")

        y = height - 160
        p.setFont("Helvetica", 12)
        p.drawString(50, y, f"Patient: {patient.get('name', appointment['patient_user_id'])}")
        y -= 20
        p.drawString(50, y, f"Doctor: {doctor.get('name', appointment['doctor_user_id'])}")
        y -= 20
        p.drawString(50, y, f"Date: {bill.get('date', '')}")
        y -= 20
        p.drawString(50, y, f"Token Number: {appointment.get('token_number', '')}")
        y -= 40

        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y, "Billing Details:")
        y -= 25
        p.setFont("Helvetica", 12)
        p.drawString(50, y, f"Appointment ID: {appointment_id}")
        y -= 20
        p.drawString(50, y, f"Amount: INR {bill.get('amount', 0)}")
        y -= 20
        p.drawString(50, y, f"Status: {bill.get('status', '')}")
        y -= 20
        p.drawString(50, y, f"Appointment Date: {appointment.get('date', '')}")
        y -= 20
        p.drawString(50, y, f"Time: {appointment.get('time', '')}")

        # Footer
        p.line(50, 120, width - 50, 120)
        p.setFont("Helvetica", 10)
        p.drawString(50, 105, hospital.get('name', 'Hospital') + ' | ' + hospital.get('location', '') if hospital else '')
        p.setFont("Helvetica", 12)
        sign_x = width - 250
        sign_y = 135
        p.drawString(sign_x, sign_y, f"Authorized Signature: {doctor.get('name', '') if doctor else ''}")
        p.line(sign_x, sign_y - 5, width - 50, sign_y - 5)

        p.showPage()
        p.save()
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"bill_{appointment_id}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        logger.error(f"Error generating bill PDF: {str(e)}")
        return jsonify({'message': 'Error generating bill'}), 500

@app.route('/patient/<token_number>', methods=['GET'])
def get_patient_by_token(token_number):
    try:
        appointments = get_db_collection('appointments')
        users = get_db_collection('users')
        appointment = appointments.find_one({'token_number': token_number})
        if not appointment:
            return jsonify({'message': 'No patient found with this token number'}), 404
        patient = users.find_one({'user_id': appointment['patient_user_id'], 'role': 'patient'})
        if not patient:
            return jsonify({'message': 'Patient not found'}), 404
        patient['_id'] = str(patient['_id'])
        patient.pop('password', None)
        patient['appointment_id'] = str(appointment['_id'])
        patient['token_number'] = appointment['token_number']
        patient['appointment_date'] = appointment['date']
        patient['appointment_time'] = appointment['time']
        doctor = users.find_one({'user_id': appointment['doctor_user_id']})
        patient['doctor_name'] = doctor['name'] if doctor else 'Unknown'
        hospital = users.find_one({'user_id': patient.get('hospital_user_id'), 'role': 'hospital'})
        patient['hospital_name'] = hospital['name'] if hospital else 'Not Assigned'
        return jsonify(patient)
    except Exception as e:
        logger.error(f"Error fetching patient by token: {str(e)}")
        return jsonify({'message': 'Error fetching patient details'}), 500

def generate_user_id(role, users):
    if role not in ['doctor', 'patient', 'hospital']:
        raise ValueError('Invalid role')
    prefix = {'doctor': 'DC_', 'patient': 'PT_', 'hospital': 'HP_'}[role]
    db = users.database
    counters = db.counters
    counters.update_one(
        {'_id': f"{role}_counter"},
        {'$setOnInsert': {'sequence_value': 1000}},
        upsert=True
    )
    result = counters.find_one_and_update(
        {'_id': f"{role}_counter"},
        {'$inc': {'sequence_value': 1}},
        return_document=True
    )  
    sequence_value = result['sequence_value']
    user_id = f"{prefix}{sequence_value}"
    if users.find_one({'user_id': user_id}):
        return generate_user_id(role, users)
    return user_id

@app.route('/hospitals', methods=['GET'])
def get_hospitals():
    try:
        users = get_db_collection('users')
        hospital_list = list(users.find({'role': 'hospital'}))
        for hospital in hospital_list:
            hospital['_id'] = str(hospital['_id'])
            hospital.pop('password', None)
        return jsonify(hospital_list)
    except Exception as e:
        logger.error(f"Error fetching hospitals: {str(e)}")
        return jsonify({'message': 'Error fetching hospitals'}), 500

@app.route('/register', methods=['POST'])
def register_user():
    users = get_db_collection('users')
    data = request.json
    role = data.get('role')
    password = data.get('password')
    name = data.get('name', '')
    specialization = data.get('specialization', '')
    phone = data.get('phone', '')
    email = data.get('email', '')
    age = data.get('age')
    gender = data.get('gender', '')
    hospital_user_id = data.get('hospital_user_id', '')
    location = data.get('location', '')

    if role == 'admin':
        return jsonify({'message': 'Cannot create admin users'}), 403
    if not password or not name:
        return jsonify({'message': 'Password and name are required'}), 400
    if role in ['doctor', 'patient'] and (not hospital_user_id or not hospital_user_id.startswith('HP_')):
        return jsonify({'message': 'Valid hospital_user_id is required for doctors and patients'}), 400

    try:
        user_id = generate_user_id(role, users)
        hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
        user_data = {
            'user_id': user_id,
            'password': hashed.decode('utf-8'),
            'role': role,
            'name': name,
            'specialization': specialization,
            'phone': phone,
            'email': email,
            'age': age,
            'gender': gender
        }
        if role in ['doctor', 'patient']:
            user_data['hospital_user_id'] = hospital_user_id
        elif role == 'hospital':
            user_data['location'] = location

        users.insert_one(user_data)
        log_activity(f"User registered: {user_id}", user_id)
        return jsonify({
            'message': 'User added successfully. Please wait for admin approval at support@netcurea.com.',
            'user_id': user_id
        })
    except ValueError:
        return jsonify({'message': 'Invalid role'}), 400
    except Exception as e:
        logger.error(f"Registration failed: {str(e)}")
        return jsonify({'message': 'Registration failed. Please contact support@netcurea.com.'}), 500

@app.route('/doctors', methods=['GET'])
def get_doctors():
    try:
        users = get_db_collection('users')
        availability = get_db_collection('availability')
        hospital_user_id = request.args.get('hospital_user_id')
        include_availability = request.args.get('include_availability', 'false').lower() == 'true'

        if not hospital_user_id or not hospital_user_id.startswith('HP_'):
            return jsonify({'message': 'Invalid hospital_user_id'}), 400

        hospital = users.find_one({'user_id': hospital_user_id, 'role': 'hospital'})
        if not hospital:
            return jsonify({'message': 'Hospital not found'}), 404

        query = {'role': 'doctor', 'hospital_user_id': hospital_user_id}  # Ensure strict filtering
        doctors = list(users.find(query))

        if not doctors:
            return jsonify({'message': 'No doctors found for this hospital', 'doctors': []}), 200

        for doc in doctors:
            doc['_id'] = str(doc['_id'])
            doc.pop('password', None)
            if include_availability:
                today = datetime.now().strftime('%Y-%m-%d')
                avail_slots = list(availability.find({
                    'doctor_user_id': doc['user_id'],
                    'date': {'$gte': today}
                }))
                for slot in avail_slots:
                    slot['_id'] = str(slot['_id'])
                doc['availability'] = avail_slots
                doc['isActive'] = len(avail_slots) > 0

        return jsonify(doctors), 200
    except Exception as e:
        logger.error(f"Error fetching doctors: {str(e)}")
        return jsonify({'message': 'Error fetching doctors'}), 500

@app.route('/staff', methods=['GET'])
def get_staff():
    users = get_db_collection('users')
    hospital_user_id = request.args.get('hospital_user_id')
    if not hospital_user_id or not hospital_user_id.startswith('HP_'):
        return jsonify({'message': 'Invalid hospital user ID'}), 400
    staff = list(users.find({'role': 'nurse', 'hospital_user_id': hospital_user_id}))
    for s in staff:
        s['_id'] = str(s['_id'])
        s.pop('password', None)
    return jsonify(staff)

@app.route('/schedule', methods=['POST'])
def add_schedule():
    users = get_db_collection('users')
    data = request.json
    hospital_user_id = data.get('hospital_user_id')
    if not hospital_user_id or not hospital_user_id.startswith('HP_'):
        return jsonify({'message': 'Invalid hospital user ID'}), 400
    schedule = {
        'hospital_user_id': hospital_user_id,
        'staff_user_id': data.get('staff_user_id'),
        'date': data.get('date'),
        'shift': data.get('shift'),
        'assigned_patient': data.get('assigned_patient'),
    }
    users.update_one(
        {'user_id': data.get('staff_user_id')},
        {'$push': {'schedule': schedule}},
        upsert=True
    )
    log_activity(f"Schedule added for staff: {data.get('staff_user_id')}", data.get('staff_user_id'))
    return jsonify({'message': 'Schedule added successfully'})

@app.route('/patients', methods=['GET'])
def get_patients():
    users = get_db_collection('users')
    hospital_user_id = request.args.get('hospital_user_id')
    if not hospital_user_id or not hospital_user_id.startswith('HP_'):
        return jsonify({'message': 'Invalid hospital user ID'}), 400
    patients = list(users.find({'role': 'patient', 'hospital_user_id': hospital_user_id}))
    for patient in patients:
        patient['_id'] = str(patient['_id'])
        patient.pop('password', None)
    return jsonify(patients), 200

@app.route('/rooms', methods=['GET'])
def get_rooms():
    rooms = get_db_collection('rooms')
    hospital_user_id = request.args.get('hospital_user_id')
    if not hospital_user_id or not hospital_user_id.startswith('HP_'):
        return jsonify({'message': 'Invalid hospital user ID'}), 400
    room_list = list(rooms.find({'hospital_user_id': hospital_user_id, 'status': 'Available'}))
    for room in room_list:
        room['_id'] = str(room['_id'])
    return jsonify(room_list), 200


@app.route('/symptoms', methods=['POST'])
def add_symptom():
    try:
        data = request.json
        user_id = data.get('user_id')
        symptoms = data.get('symptoms', [])
        details = data.get('details', {})  # New field for detailed data
        notes = data.get('notes', '')

        if not user_id or not user_id.startswith('PT_') or not symptoms:
            return jsonify({'message': 'Invalid input'}), 400

        symptom_entry = {
            'user_id': user_id,
            'symptoms': symptoms,
            'details': details,  # e.g., {'fever': {'temperature': 38.5}, 'stomach_pain': {'location': 2, 'intensity': 7}}
            'notes': notes,
            'created_at': datetime.utcnow().isoformat(),
            'status': 'pending'
        }
        get_db_collection('symptoms').insert_one(symptom_entry)
        return jsonify({'message': 'Symptoms recorded successfully', 'symptom_id': str(symptom_entry['_id'])}), 201
    except Exception as e:
        logger.error(f"Error adding symptom: {str(e)}")
        return jsonify({'message': 'Error recording symptoms'}), 500

@app.route('/symptoms/analyze', methods=['POST'])
def analyze_symptoms():
    try:
        data = request.json
        symptoms = data.get('symptoms', [])
        details = data.get('details', {})

        if not symptoms:
            return jsonify({'message': 'No symptoms provided'}), 400

        # Simple rule-based analysis with details
        if 'fever' in symptoms and 'temperature' in details.get('fever', {}):
            temp = details['fever']['temperature']
            if temp > 38.0:
                return jsonify({
                    'symptoms': symptoms,
                    'predicted_diagnosis': 'Possible High Fever (Influenza or Infection)',
                    'confidence': 0.8,
                    'recommendation': 'Consult a doctor immediately.'
                })
            else:
                return jsonify({
                    'symptoms': symptoms,
                    'predicted_diagnosis': 'Mild Fever',
                    'confidence': 0.6,
                    'recommendation': 'Monitor and consult if persists.'
                })

        if 'stomach_pain' in symptoms and 'location' in details.get('stomach_pain', {}) and 'intensity' in details.get('stomach_pain', {}):
            location = details['stomach_pain']['location']  # 1: Upper, 2: Middle, 3: Lower
            intensity = details['stomach_pain']['intensity']  # 1-10 scale
            if intensity > 7 and location == 3:
                return jsonify({
                    'symptoms': symptoms,
                    'predicted_diagnosis': 'Possible Appendicitis',
                    'confidence': 0.7,
                    'recommendation': 'Seek emergency care.'
                })
            else:
                return jsonify({
                    'symptoms': symptoms,
                    'predicted_diagnosis': 'General Stomach Discomfort',
                    'confidence': 0.5,
                    'recommendation': 'Consult a doctor.'
                })

        return jsonify({
            'symptoms': symptoms,
            'predicted_diagnosis': 'Unknown Condition',
            'confidence': 0.3,
            'recommendation': 'Please consult a doctor for detailed assessment.'
        }), 200
    except Exception as e:
        logger.error(f"Error analyzing symptoms: {str(e)}")
        return jsonify({'message': 'Error analyzing symptoms'}), 500
if __name__ == '__main__':
    app.run(debug=True)